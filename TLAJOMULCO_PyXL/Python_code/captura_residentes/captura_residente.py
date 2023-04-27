# LIBRERIAS
import json
import openpyxl
import xlwings as xw


# MODULOS


# VARIABLES GLOBALES


# FUNCIONES

def extraccion_datos_reporte():

    # Definimos la ruta del reporte desde donde extraeremos la informacion
    ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\TLAJOMULCO_PyXL\\Python_code\\captura_residentes\\EJEMPLO_REPORTE.xlsx"

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    # Definimos la lista que llevara los diccionarios para trabajar mas adelante con esta estructura de datos
    datos_reporte = []
    
    # Definimos el primer valor como None para la asignacion inicial
    codigo_de_nomina = None

    # Ciclo For que comenzara encontrando en nuestro reporte la celda con "CODIGO" en la columna A para iniciar la extraccion de valores
    for fila in hoja['A']:
        if fila.value == "CODIGO":
            # Asignamos el valor 
            codigo_de_nomina = hoja.cell(row=fila.row+1, column=1).value

            # El ciclo se rompera cuando no se encuentre ningun valor bajo la celda "CODIGO", lo que significa que el reporte no tiene mas informacion
            if codigo_de_nomina is None:
                break
            
            # La excepcion de la condicion es cuando se encuentra un valor y comenzamos con la extracion de informacion
            else:
                # Nombre del residente - REPRESENTADO EN EL ARCHIVO EXCEL
                nombre = hoja.cell(row=fila.row+1, column=2).value

                # List Comprenhension, algo compleja ... para extraer y concatenar los dias y su distribucion en las obras
                distribucion_obras = [[[hoja.cell(row=fila.row+i, column=4).value], [hoja.cell(row=fila.row+i, column=5).value]] for i in range(2, 8) if hoja.cell(row=fila.row+i, column=4).value is not None and hoja.cell(row=fila.row+i, column=5).value is not None]
                
                # Despues de tener una lista de varias listas, se concatena de esta forma los dias en las obras que se repiten, para la asignacion de valores
                result_dict = {}
                for sublist in distribucion_obras:
                    key = sublist[0][0]
                    value = sublist[1][0]
                    if key in result_dict:
                        result_dict[key] += value
                    else:
                        result_dict[key] = value

                # Simplemente convertimos nuestra variable resultados de diccionario en nuestra nueva distribucion de obras
                distribucion_obras = result_dict

                # Si encontramos valor en domingo, lo asignaremos como Domingo trabajado
                domingo_trabajado = {}

                key = hoja.cell(row=fila.row+8, column=4).value
                value = hoja.cell(row=fila.row+8, column=5).value

                if key is not None and value is not None:
                    domingo_trabajado[key] = value


                # Asignacion final con el metodo append a nuestra lista llamada datos_reporte, la cual contendra todos los datos de los trabajadores
                if len(domingo_trabajado) > 0:
                    datos_reporte.append({"codigo_de_nomina": codigo_de_nomina, "nombre": nombre, "distribucion_obras": distribucion_obras, "domingo_trabajado": domingo_trabajado})
                else:
                    datos_reporte.append({"codigo_de_nomina": codigo_de_nomina, "nombre": nombre, "distribucion_obras": distribucion_obras})

    
    return datos_reporte

datos_reporte = extraccion_datos_reporte()
print(json.dumps(datos_reporte, indent=4))

def celda_inicio_residentes():
    pass
