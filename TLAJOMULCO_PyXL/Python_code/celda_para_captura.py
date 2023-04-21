# LIBRERIAS
import openpyxl

def obtener_celda_para_captura(archivo_para_captura):

    # Cargamos el archivo_para_captura de Excel
    wb = openpyxl.load_workbook(archivo_para_captura)
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.active

    # Inicializar las variables que almacenarán la celda con el último valor encontrado
    ultima_fila_b = 0
    ultima_fila_d = 0

    # Recorrer las filas del rango especificado
    for fila in range(14, 301):

        # Obtener el valor de la columna B en la fila actual
        valor_b = ws.cell(row=fila, column=2).value
        # Si el valor es un número menor a 70, lo almacenamos
        if isinstance(valor_b, (int, float)) and valor_b < 70:
            ultima_fila_b = fila

        # Obtener el valor de la columna D en la fila actual
        valor_d = ws.cell(row=fila, column=4).value
        # Si el valor es un string, lo almacenamos
        if isinstance(valor_d, str):
            ultima_fila_d = fila
        
    # Obtener la celda para captura

    # Si la fila de b y d son 0, retornaremos la celda inicial
    if ultima_fila_b == 0 and ultima_fila_d == 0:
        celda_para_captura = ws.cell(row=15, column=1)
    # Si la fila b es igual que fila b, retornamos celda_captura
    elif ultima_fila_b == ultima_fila_d:
        celda_para_captura = ws.cell(row=ultima_fila_b +2, column=1)
    # Si la fila b es mayor que fila b, retornamos celda_captura
    elif ultima_fila_b > ultima_fila_d:
        celda_para_captura = ws.cell(row=ultima_fila_b +2, column=1)
    else:
        celda_para_captura = ws.cell(row=ultima_fila_d +2, column=1)

    # Obetener el ultimo valor

    # Empezamos a buscar desde la fila 14
    fila_actual = 14
        
    # Inicializamos el valor a devolver con None
    ultimo_valor = None
        
    # Recorremos todas las filas de la hoja hasta encontrar un valor numérico menor a 70
    while fila_actual <= 300:
        celda_b = ws.cell(row=fila_actual, column=2)
        valor_b = celda_b.value
            
        # Si la celda B de la fila actual tiene un valor numérico, lo guardamos como último valor
        if isinstance(valor_b, (int, float)):
            if valor_b < 70 and (ultimo_valor is None or valor_b > ultimo_valor):
                ultimo_valor = valor_b
            
        fila_actual += 1
        
    # Si no se encontró ningún valor menor a 70, se devuelve 1
    if ultimo_valor is None:
        ultimo_valor = 1
    else:
        ultimo_valor += 1

    return celda_para_captura, ultimo_valor

