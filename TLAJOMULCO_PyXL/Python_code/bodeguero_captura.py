# LIBRERIAS
import openpyxl
import os
import xlwings as xw

# MODULOS
from celda_para_captura import obtener_celda_para_captura

# VARIABLES GLOBALES - ASIGNADAS POR EL USUARIO EN CONSOLA
bodeguero = str(input("¿A que bodeguero piensas capturar?: "))
numero_de_semana = int(input("Ingrese el numero de semana que estara trabajando: "))
ruta_carpeta_semana = f"C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\TLAJOMULCO_PyXL\\SEMANA_{numero_de_semana}"
dias_bodeguero = int(input("¿Cuantos dias trabajo el bodeguero?: "))

while True:
    vacaciones_bodeguero = input("¿El bodeguero tomó vacaciones? (Y/N): ").upper()
    if vacaciones_bodeguero == "Y":
        break
    elif vacaciones_bodeguero == "N":
        dias_bodeguero = (dias_bodeguero / 6) * 7
        break
    else:
        print("Entrada no válida. Por favor, ingrese 'Y' o 'N'.")


def obtener_total_destajo():

    subtotales_de_obras = []
    for filename in os.listdir(ruta_carpeta_semana):
        if filename.endswith(".xlsm"):
            filepath = os.path.join(ruta_carpeta_semana, filename)
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active

            # Verificar si el valor de la celda B10 es igual al número de semana
            if ws["B10"].value == numero_de_semana:
                 # Recorrer la columna O desde la fila 15 hasta encontrar la celda que tenga el valor de "SUBTOTAL"
                for i in range(15, ws.max_row):
                    if ws.cell(row=i, column=15).value == "SUBTOTAL":
                        # Almacenar el valor de la celda +2 columnas a la derecha en un diccionario
                        nombre_archivo = filename.split(" ")[0]
                        subtotal = ws.cell(row=i, column=15+2).value
                        subtotales_de_obras.append({nombre_archivo: subtotal})
                        break

    return subtotales_de_obras

subtotales_de_obras = obtener_total_destajo()

def obtener_prorrateo_bodeguero():
    total_destajos = 0

    for subtotal in subtotales_de_obras:
        subtotal_values = subtotal.values()
        subtotal_total = sum(subtotal_values)
        total_destajos += subtotal_total

    total_destajos = round(total_destajos, 2)

    prorrateo_bodeguero = []
    dias_bodega = dias_bodeguero

    for subtotal in subtotales_de_obras:
        prorrateo = {}
        for key, value in subtotal.items():
            prorrateo[key] = round((value / total_destajos) * dias_bodega, 3)
        prorrateo_bodeguero.append(prorrateo)

    prorrateo_bodeguero = [[key, value] for diccionario in prorrateo_bodeguero for key, value in diccionario.items()]

    return prorrateo_bodeguero

prorrateo_bodeguero = obtener_prorrateo_bodeguero()

def obtener_datos_bodeguero():

    ruta_carpeta_semana = f"C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\TLAJOMULCO_PyXL\\SEMANA_{numero_de_semana}\\"

    datos_para_capturar_bodeguero = prorrateo_bodeguero
    for i in range(len(datos_para_capturar_bodeguero)):
        original_valor = datos_para_capturar_bodeguero[i][0]
        archivos_en_directorio = os.listdir(ruta_carpeta_semana)
        for archivo in archivos_en_directorio:
            if archivo.startswith(original_valor.split()[0]):
                datos_para_capturar_bodeguero[i][0] = ruta_carpeta_semana + archivo
                break
    

    datos_para_capturar_bodeguero
    for obra in datos_para_capturar_bodeguero:
        celda_para_capturar, ultimo_valor = obtener_celda_para_captura(obra[0])
        obra.append(celda_para_capturar)
        obra.append(ultimo_valor)

    return datos_para_capturar_bodeguero

datos_para_captura_bodeguero = obtener_datos_bodeguero()

def captura_bodeguero(sublista):

    obra_para_captura = 0
    dia_prorrateado = 1
    celda_para_captura = 2
    ultimo_valor = 3

    # Cargamos el archivo_para_captura de Excel
    wb = xw.Book(sublista[obra_para_captura])
    # Seleccionamos la hoja en la que queremos buscar
    ws = wb.sheets.active

    # Tomamos el valor de la ultima celda de captura
    fila = sublista[celda_para_captura].row
    columna = sublista[celda_para_captura].column

    # Asignar el codigo de nomina del trabajador
    codigo = ws.range(fila, columna).value = bodeguero
    
    # Asignar los valores prorrateados
    celda_orden1 = ws.range(fila, columna +1).value = sublista[ultimo_valor]
    celda_orden2 = ws.range(fila, columna +15).value = sublista[ultimo_valor]
    celda_dias = ws.range(fila, columna +11).value = sublista[dia_prorrateado]
    celda_porcentaje = ws.range(fila, columna +18).value = 1
    celda_actividades = ws.range(fila +1, columna +3).value = "Bodeguero, recibir y entregar materiales"

    ruta = sublista[obra_para_captura]
    separador = '\\'

    # dividir la ruta en una lista
    ruta_lista = ruta.split(separador)

    # encontrar el índice del elemento que contiene "SEMANA_14"
    indice_semana = ruta_lista.index(f'SEMANA_{numero_de_semana}')

    # acceder al siguiente elemento de la lista, que es el nombre del archivo
    nombre_archivo = ruta_lista[indice_semana + 1]

    # dividir el nombre del archivo en una lista
    nombre_archivo_lista = nombre_archivo.split('.')

    # acceder al primer elemento de la lista, que es el nombre de la obra
    obra = nombre_archivo_lista[0]

    # Guardar y cerrar el libro de Excel
    wb.save(sublista[obra_para_captura])
    
    return print("Se ha capturado al bodeguero en la obra", obra, "En la celda:", sublista[celda_para_captura].coordinate)

for sublista in datos_para_captura_bodeguero:
    captura_bodeguero(sublista)

print("SE TERMINO LA CAPTURA DEL BODEGUERO")









    
    

