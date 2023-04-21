# LIBRERIAS 
import openpyxl
import xlwings as xw

# MODULOS
from datos_de_captura import datos_de_captura
from archivo_para_captura import obtener_archivo_para_captura

def buscar_celda_mandos_intermedios(clave):

    archivo_para_captura = obtener_archivo_para_captura(clave)
    libro = openpyxl.load_workbook(archivo_para_captura)
    hoja = libro.active

    # Buscar la celda que contiene el valor "MANDOS INTERMEDIOS"
    for fila in range(15, hoja.max_row + 1):
        celda = hoja.cell(row=fila, column=7)
        if celda.value == "MANDOS INTERMEDIOS":
            # Si se encuentra el valor, devolver la celda dos filas más abajo y seis columnas menos
            celda_mandos_intermedios = hoja.cell(row=fila + 3, column=celda.column - 6)
            return celda_mandos_intermedios

def valor_total_destajo(clave):

    archivo_para_captura = obtener_archivo_para_captura(clave)
    libro = openpyxl.load_workbook(archivo_para_captura, data_only=True)
    hoja = libro.active

    # Buscar la celda que contiene el total del destajo
    for fila in range(15, hoja.max_row +1):
        celda = hoja.cell(row=fila, column=15)
        if celda.value == "SUBTOTAL":
            # Si se encuentra el valor, devolver la celda dos columnas a la derecha
            total_destajo = hoja.cell(row=fila, column=celda.column + 2).value

            return '$ {:,.3f}'.format(round(float(total_destajo), 3))

def capturar_mandos_intermedios(clave):

    # VARIABLES DE LA FUNCIONS
    archivo_para_captura = obtener_archivo_para_captura(clave)
    celda_mandos_intermedios = buscar_celda_mandos_intermedios(clave)
    total_destajo = valor_total_destajo(clave)
    obra = 1

    # PORCENTAJE DE MANDOS INTERMEDIOS
    porcentaje_mandos_intermedios = input(f"Ingrese el porcentaje de mandos intermedios en la obra {datos_de_captura[clave][obra]} : ")
    porcentaje_elegido = float(porcentaje_mandos_intermedios) /100

    if celda_mandos_intermedios.value == 'lote':
        pass
    else:
        # Cargamos el archivo_para_captura de Excel
        wb = xw.Book(archivo_para_captura)
        # Seleccionamos la hoja en la que queremos buscar
        ws = wb.sheets.active

        # Tomamos el valor de la ultima celda de captura
        fila = celda_mandos_intermedios.row
        columna = celda_mandos_intermedios.column

        # Asignar el codigo
        codigo = ws.range(fila, columna).value = "lote"
        #Asignar ordenes
        orden1 = ws.range(fila, columna +1).value = 80
        orden2 = ws.range(fila, columna +15).value = 80

        #Asignar concepto
        mandos1 = ws.range(fila, columna +7).value = porcentaje_elegido
        mandos2 = ws.range(fila, columna +11).value = 1
        mandos3 = ws.range(fila, columna +3).value = f"MANDOS INTERMEDIOS  %{porcentaje_mandos_intermedios} DE {total_destajo}"
        if ws.range(fila, columna +18).value is not None:
            ws.range(fila, columna +18).clear_contents()
        ws.range(fila, columna + 17).formula = "=RC[-1]"

        # Guardar y cerrar el archivo de Excel
        wb.save(archivo_para_captura)
        wb.close()
        
        return print(f"Se capturaron los mandos intermedios de la obra:  {datos_de_captura[clave][obra]}")
    

