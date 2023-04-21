# LIBRERIAS 
import openpyxl
import xlwings as xw

# MODULOS
from datos_de_captura import datos_de_captura
from archivo_para_captura import obtener_archivo_para_captura

def buscar_celda_en_archivo(clave):

    archivo_para_captura = obtener_archivo_para_captura(clave)
    libro = openpyxl.load_workbook(archivo_para_captura)
    hoja = libro.active

    # Buscar la celda que contiene el valor "RESIDENTES, SUPERVISORES Y DESTAJISTAS DE OBRA"
    for fila in range(15, hoja.max_row + 1):
        celda = hoja.cell(row=fila, column=7)
        if celda.value == "RESIDENTES, SUPERVISORES Y DESTAJISTAS DE OBRA":
            # Si se encuentra el valor, devolver la celda dos filas más abajo y seis columnas menos
            celda_destajista = hoja.cell(row=fila + 3, column=celda.column - 6)
            return celda_destajista

def capturar_destajista(clave):

    # VARIABLES DE LA FUNCION
    archivo_para_captura = obtener_archivo_para_captura(clave)
    celda_destajista = buscar_celda_en_archivo(clave)
    
    if celda_destajista.value == 34:
        pass
    else:
        # Cargamos el archivo_para_captura de Excel
        wb = xw.Book(archivo_para_captura)
        # Seleccionamos la hoja en la que queremos buscar
        ws = wb.sheets.active

        # Tomamos el valor de la ultima celda de captura
        fila = celda_destajista.row
        columna = celda_destajista.column

        # Asignar el codigo
        codigo = ws.range(fila, columna).value = 34
        #Asignar ordenes
        orden1 = ws.range(fila, columna +1).value = 71
        orden2 = ws.range(fila +1, columna +15).value = 71
        orden3 = ws.range(fila + 1, columna +1).value = 71
        #Asignar porcentaje
        porcentaje = ws.range(fila, columna +18).value = 1
        #Asignar concepto
        destajo = ws.range(fila + 1, columna).value = "destaj"
        destajo1 = ws.range(fila + 1, columna+7).value = 0.04
        destajo2 = ws.range(fila + 1, columna+11).value = 1

        wb.save(archivo_para_captura)
        return print(f"Se ha capturado al destajista en la obra {datos_de_captura[clave][1]}")


