# LIBRERIAS
import openpyxl
import datetime
import locale

# Configurar el locale en español // PARA LA CONVERSION DE FECHAS
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

# VARIABLES
ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\JUEVES_PyXl\\JUEVES_REPORTE.xlsx"

def obtener_datos_de_captura(ruta_archivo_origen):

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    # Definimos el valor del domingo que fue trabajado en esta semana
    fecha_domingo = str(hoja.cell(row=3, column=10).value)
    fecha_domingo = datetime.datetime.strptime(fecha_domingo, '%Y-%m-%d %H:%M:%S')
    fecha_domingo -= datetime.timedelta(days=1)
    fecha_domingo = fecha_domingo.strftime('%d de %B')

    # Si hay dia festivo en la semana, se asigna la fecha en el reporte y en las variables
    fecha_dia_festivo = str(hoja.cell(row=3, column=12).value)
    fecha_dia_festivo = datetime.datetime.strptime(fecha_dia_festivo, '%Y-%m-%d %H:%M:%S')
    fecha_dia_festivo = fecha_dia_festivo.strftime('%d de %B')

    # Definimos la lista donde vamos a guardar los datos
    datos_de_captura = []

    # Empezamos a leer desde la fila 5
    fila = 5

    # Iteramos mientras haya datos en la columna A
    while hoja.cell(row=fila, column=1).value:
        """Extraemos los valores de las columnas A, C, D, E, F, G, I, J, K
        Los cuales representaran Codigo, Obra, Dias, Tiempo extra, Domingo, Vacaciones, Actividades, Velador, Dia festivo y Salario"""

        valores_fila = [hoja.cell(row=fila, column=columna).value for columna in range(1, 12) if columna in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11]]

        # Convertimos el valor de la celda I a un string
        if isinstance(valores_fila[8], float) or isinstance(valores_fila[8], int):
            valores_fila[8] = str(valores_fila[8])

        # AGREGAMOS EL IF STATEMENT PARA COMPROBAR LAS VACACIONES
        if valores_fila[5] is None:
            # Multiplicamos el valor de la columna D = Dias por 7/6
            valores_fila[2] = valores_fila[2] * 7/6
        else:
            valores_fila[2]
        
        # Convertimos los dias velador a proporcional
        if valores_fila[5] is None:
            if valores_fila[7] is None:
                pass
            else:
                valores_fila[7] = valores_fila[7] * 7/6 # Multiplicamos el valor de la columna D = Dias por 7/6
        else:
            valores_fila[7]

        # Agregamos los valores a la lista de datos
        datos_de_captura.append(valores_fila)

        # Avanzamos a la siguiente fila
        fila += 1

    # GENERAMOS LA LISTA DE ACTIVIDADES
    # Creamos una sublista que representa las actividades de cada trabajador
    actividades = []
    for sublista in datos_de_captura:
        actividad = sublista[6]
        if actividad is None or actividad == "":
            actividad = "" # Si actividad es None o una cadena vacía, asignamos una cadena vacía
        actividades.append(actividad)

    lista_de_actividades = [[] for i in range(len(actividades))]

    for i, actividad in enumerate(actividades):
        # Dividir la cadena de texto en subcadenas de máximo 46 caracteres
        subcadenas = []
        while len(actividad) > 0:
            if len(actividad) <= 46:
                subcadenas.append(actividad)
                actividad = ""
            else:
                espacio = actividad.rfind(" ", 0, 46)
                if espacio == -1:
                    subcadenas.append(actividad[:46])
                    actividad = actividad[46:]
                else:
                    subcadenas.append(actividad[:espacio])
                    actividad = actividad[espacio+1:]

        # Agregar las subcadenas a la nueva lista correspondiente
        lista_de_actividades[i].extend(subcadenas)

    # GENERAMOS LA LISTA DE TRABAJADORES
    # La lista trabajadores, contendra las claves de cada uno de los trabajadores en datos_de_captura
    trabajadores = [lista[0] for lista in datos_de_captura]

    # Este ciclo for llenara la lista trabajador, enumerando a trabajadores iniciando desde el 0, para poder usarla como parametro en nuestra variable
    # Enumeramos los elementos de la lista y guardamos los índices en una lista
    trabajador = [i for i, _ in enumerate(trabajadores)]

    return datos_de_captura, lista_de_actividades, trabajador, fecha_domingo, fecha_dia_festivo

datos_de_captura, lista_de_actividades, trabajador, fecha_domingo, fecha_dia_festivo = obtener_datos_de_captura(ruta_archivo_origen)

