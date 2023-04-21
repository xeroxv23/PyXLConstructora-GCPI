# LIBRERIAS
import openpyxl
import datetime
import locale

# Configurar el locale en español // PARA LA CONVERSION DE FECHAS
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

# VARIABLES
ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\LAPLANTA_PyXL\\ZONA_INDUSTRIAL.xlsx"

def obtener_datos_de_captura(ruta_archivo_origen):

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    # Definimos la fecha de los dias de la semana
    
    # Crear el diccionario
    fecha_dia_semana = {}

    # Agregar las claves y valores al diccionario
    fecha_dia_semana['LUNES'] = hoja.cell(row=4, column=4).value.strftime("%d de %B")
    fecha_dia_semana['MARTES'] = hoja.cell(row=4, column=9).value.strftime("%d de %B")
    fecha_dia_semana['MIERCOLES'] = hoja.cell(row=4, column=14).value.strftime("%d de %B")
    fecha_dia_semana['JUEVES'] = hoja.cell(row=4, column=19).value.strftime("%d de %B")
    fecha_dia_semana['VIERNES'] = hoja.cell(row=4, column=24).value.strftime("%d de %B")
    fecha_dia_semana['SABADO'] = hoja.cell(row=4, column=29).value.strftime("%d de %B")
    fecha_dia_semana['DOMINGO'] = hoja.cell(row=4, column=34).value.strftime("%d de %B")

    # Si hay dia festivo en la semana, se asigna la fecha en el reporte y en las variables
    dia_festivo_semana = hoja.cell(row=4, column=38).value.strftime("%d de %B")

    # Definimos la lista donde vamos a guardar los datos
    datos_de_captura = []

    # Empezamos a leer desde la fila 5
    fila = 6

    # ITERAMOS ESTE CODIGO MIENTRAS EL REPORTE TENGA DATOS
    while hoja.cell(row=fila, column=1).value:
        """ EXTRAEMOS TODOS LOS VALORES DE LAS CELDAS PARA ASGINACION DE DATOS DE CAPTURA"""

        valores_fila = [hoja.cell(row=fila, column=columna).value for columna in range(1, 41) if columna != 2]

        # CONVERTIR EL VALOR DE LOS INDICES DE ACTIVIDADES A STRING
        posiciones_a_convertir = [6, 11, 16, 21, 26, 31, 36]

        for posicion in posiciones_a_convertir:
            if isinstance(valores_fila[posicion], (float, int)):
                valores_fila[posicion] = str(valores_fila[posicion])

        # SI EXISTE ALGUN VALOR == 0 ( QUE ES EL CODIGO PARA ASIGNAR VACACIONES) LOS VALORE SE MANTENDRAN SIN PROPORCIONAL , DE LO CONTRARIO SE MULTIPLICARAN POR 7/6, SE CAMBIA EL VALOR 0 AL STRING " VACACIONES " 
        indice_dias_semana = [2, 7, 12, 17, 22, 27]

        if 1 in [valores_fila[i] for i in indice_dias_semana] or None in [valores_fila[i] for i in indice_dias_semana]:
            if 0 not in [valores_fila[i] for i in indice_dias_semana]:
                for i in indice_dias_semana:
                    if isinstance(valores_fila[i], (int, float)) and valores_fila[i] in [1, None] or isinstance(valores_fila[i], float):
                        valores_fila[i] = valores_fila[i] * 7/6
            elif 0 in [valores_fila[i] for i in indice_dias_semana]:
                for i in indice_dias_semana:
                    if valores_fila[i] == 0:
                        valores_fila[i] = "Vacaciones"

        # Agregamos los valores a la lista de datos
        datos_de_captura.append(valores_fila)

        # Avanzamos a la siguiente fila
        fila += 1
        return datos_de_captura, fecha_dia_semana, dia_festivo_semana

datos_de_captura, fecha_dia_semana, dia_festivo_semana = obtener_datos_de_captura(ruta_archivo_origen)
print(datos_de_captura)

"""# GENERAMOS LA LISTA DE ACTIVIDADES
    # Creamos una sublista que representa las actividades de cada trabajador
    actividades = []
    for sublista in datos_de_captura:
        actividad = sublista[6]
        if actividad is None or actividad == "":
            actividad = "" # Si actividad es None o una cadena vacía, asignamos una cadena vacía
        actividades.append(actividad)

    lista_de_actividades = [[] for i in range(len(actividades))]

    for i, actividad in enumerate(actividades):
        # Dividir la cadena de texto en subcadenas de máximo 46 caracteres
        subcadenas = []
        while len(actividad) > 0:
            if len(actividad) <= 46:
                subcadenas.append(actividad)
                actividad = ""
            else:
                espacio = actividad.rfind(" ", 0, 46)
                if espacio == -1:
                    subcadenas.append(actividad[:46])
                    actividad = actividad[46:]
                else:
                    subcadenas.append(actividad[:espacio])
                    actividad = actividad[espacio+1:]

        # Agregar las subcadenas a la nueva lista correspondiente
        lista_de_actividades[i].extend(subcadenas)

    # GENERAMOS LA LISTA DE TRABAJADORES
    # La lista trabajadores, contendra las claves de cada uno de los trabajadores en datos_de_captura
    trabajadores = [lista[0] for lista in datos_de_captura]

    # Este ciclo for llenara la lista trabajador, enumerando a trabajadores iniciando desde el 0, para poder usarla como parametro en nuestra variable
    # Enumeramos los elementos de la lista y guardamos los índices en una lista
    trabajador = [i for i, _ in enumerate(trabajadores)]

    return datos_de_captura, lista_de_actividades, trabajador, fecha_domingo, fecha_dia_festivo

datos_de_captura, lista_de_actividades, trabajador, fecha_domingo, fecha_dia_festivo = obtener_datos_de_captura(ruta_archivo_origen)
"""
