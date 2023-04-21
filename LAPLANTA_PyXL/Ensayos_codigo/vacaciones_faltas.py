# LIBRERIAS
import openpyxl
import locale

# Configurar el locale en español // PARA LA CONVERSION DE FECHAS
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

def obtener_los_trabajadores():

    ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\LAPLANTA_PyXL\\EJEMPLO_REPORTE.xlsx"

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    # COMENZAREMSO CON EL CODIGO DE NOMINA
    trabajadores = []

    # Busca desde la fila 6 en la columna 1 hasta el final de la columna
    for row in hoja.iter_rows(min_row=6, min_col=1):
        # Comprueba si la celda contiene el string "CODIGO"
        if row[0].value == "CODIGO":
            # Si la siguiente fila no es None, agrega el valor a la lista
            if row[0].offset(row=1).value is not None:
                trabajadores.append(row[0].offset(row=1).value)
    
    return trabajadores

trabajadores = obtener_los_trabajadores()

def obtener_vacaciones_faltas_trabajadores(codigo_de_nomina):

    lista_vacaciones = []
    lista_faltas = []
    codigo_de_nomina = codigo_de_nomina
    ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\LAPLANTA_PyXL\\EJEMPLO_REPORTE.xlsx"

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    # Busca desde la fila 6 en la columna 1 hasta el final de la columna
    for row in hoja.iter_rows(min_row=6, min_col=1):
        # Comprueba si la celda contiene el string "CODIGO"
        if row[0].value == codigo_de_nomina:
                
            # Obtiene los datos estáticos que no serán modificados por días
            codigo_de_nomina = hoja.cell(row=row[0].row, column=row[0].column).value
            nombre = hoja.cell(row=row[0].row, column=row[0].column+1).value
            
            # Iterar sobre los dias de la semana para obtener las vacaciones
            for i in range(1,7):
                vacaciones_tomadas = hoja.cell(row=row[0].row+i, column=row[0].column+8).value
                if vacaciones_tomadas == 1:
                    fecha_de_vacaciones_tomadas = hoja.cell(row=row[0].row+i, column=row[0].column+2).value.strftime("%d de %B del %Y")
                    lista_vacaciones.append(fecha_de_vacaciones_tomadas)
            
            # Iterar sobre los dias de la semana para obtener las vacaciones
            for i in range(1,7):
                fecha_de_falta = hoja.cell(row=row[0].row+i, column=row[0].column+4).value
                vacaciones_tomadas = hoja.cell(row=row[0].row+i, column=row[0].column+8).value
                if fecha_de_falta is None and vacaciones_tomadas is None:
                    fecha_de_falta = hoja.cell(row=row[0].row+i, column=row[0].column+2).value.strftime("%d de %B del %Y")
                    lista_faltas.append(fecha_de_falta)
            
            if len(lista_vacaciones) >= 1:
                lista_vacaciones.insert(0, "Vacaciones tomadas")
                lista_vacaciones.insert(1, codigo_de_nomina)
                lista_vacaciones.insert(2, nombre)
            if len(lista_faltas) >= 1:
                lista_faltas.insert(0, "Faltas")
                lista_faltas.insert(1, codigo_de_nomina)
                lista_faltas.insert(2, nombre)

    return lista_vacaciones, lista_faltas

for trabajador in trabajadores:

    lista_vacaciones, lista_faltas = obtener_vacaciones_faltas_trabajadores(trabajador)
    if len(lista_vacaciones) == 0:
        lista_vacaciones = None
    else:
        print(lista_vacaciones)
    if len(lista_faltas) == 0:
        lista_faltas = None
    else:
        print(lista_faltas)
                        
        

    