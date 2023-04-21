import openpyxl

def obtener_datos_de_captura_por_obra(codigo_de_nomina, obra):

    # VARIABLES
    ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\LAPLANTA_PyXL\\EJEMPLO_REPORTE.xlsx"

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    celda_inicio = None
    datos_de_captura = {}

    # Busca desde la fila 6 en la columna 1 hasta el final de la columna
    for row in hoja.iter_rows(min_row=6, min_col=1):
        # Comprueba si la celda contiene el string "CODIGO"
        if row[0].value == codigo_de_nomina:
            celda_inicio = row[0]
            datos_de_captura["codigo"] = codigo_de_nomina
            salario = hoja.cell(row=row[0].row+3, column=row[0].column).value
            datos_de_captura["salario"] = salario
            datos_de_captura["obra"] = obra
        
        # Haremos la comprobacion de las celdas para ver si agregamos el set de datos al diccionario
            dias_semana = ['lunes', 'martes', 'miercoles', 'jueves', 'viernes', 'sabado', 'domingo']
            fila_inicio = celda_inicio.row
            columna_inicio = celda_inicio.column
            valores_celdas = [hoja.cell(fila_inicio+i, columna_inicio+3).value for i in range(1, 8)]

            for i, dia in enumerate(dias_semana):
                if valores_celdas[i] == obra:
                    valores_dia = [hoja.cell(fila_inicio+i+1, columna_inicio+j+4).value for j in range(9) if j+1 in range(1, 9)]
                    if any(valores_dia):
                        datos_de_captura[f"valores_{dia}"] = valores_dia

    return datos_de_captura

datos_de_captura = obtener_datos_de_captura_por_obra(1576, "A-002")
print(datos_de_captura)