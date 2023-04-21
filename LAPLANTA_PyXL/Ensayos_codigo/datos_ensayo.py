# LIBRERIAS
import openpyxl
import datetime
import locale

# Configurar el locale en español // PARA LA CONVERSION DE FECHAS
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')

# VARIABLES
ruta_archivo_origen = "C:\\Users\\carlo\\Google Drive\\GCPI - TRABAJO\\PyXLConstruc\\LAPLANTA_PyXL\\EJEMPLO_REPORTE.xlsx"

def obtener_datos_de_captura(ruta_archivo_origen):

    # Cargamos el archivo de Excel con openpyxl
    libro = openpyxl.load_workbook(ruta_archivo_origen, data_only=True)
    hoja = libro.active

    # Definimos la fecha de los dias de la semana
    
    # Crear el diccionario
    fecha_dia_semana = {}

    # Agregar las claves y valores al diccionario
    fecha_dia_semana['LUNES'] = hoja.cell(row=9, column=3).value.strftime("%d de %B")
    fecha_dia_semana['MARTES'] = hoja.cell(row=10, column=3).value.strftime("%d de %B")
    fecha_dia_semana['MIERCOLES'] = hoja.cell(row=11, column=3).value.strftime("%d de %B")
    fecha_dia_semana['JUEVES'] = hoja.cell(row=12, column=3).value.strftime("%d de %B")
    fecha_dia_semana['VIERNES'] = hoja.cell(row=13, column=3).value.strftime("%d de %B")
    fecha_dia_semana['SABADO'] = hoja.cell(row=14, column=3).value.strftime("%d de %B")
    fecha_dia_semana['DOMINGO'] = hoja.cell(row=15, column=3).value.strftime("%d de %B")

    """ OBTENDREMOS LOS DATOS PARA CAPTURA, MIENTRAS EXISTA UN CELDA CON CODIGO DE NOMINA, BAJO LA CABECERA CODIGO EN EL ARCHIVO DE REPORTE
     DE AHI TOMAREMOS COMO REFERENCIA LOS DATOS QUE SE CAPTURARAN EN DESTAJOS """
    
    # COMENZAREMSO CON EL CODIGO DE NOMINA
    codigo_de_nomina = []

    # Busca desde la fila 6 en la columna 1 hasta el final de la columna
    for row in hoja.iter_rows(min_row=6, min_col=1):
        # Comprueba si la celda contiene el string "CODIGO"
        if row[0].value == "CODIGO":
            # Si la siguiente fila no es None, agrega el valor a la lista
            if row[0].offset(row=1).value is not None:
                codigo_de_nomina.append(row[0].offset(row=1).value)

    # LOS SIGUIENTES SON LOS DATOS DE CAPTURA, LOS CUALES TENEMOS QUE VER LA FORMA DE SINTETIZARLOS
    # Crea una lista vacía para almacenar los datos de captura del trabajador
    datos_de_captura_trabajador = []

    # Busca los valores de la lista en la columna 1 de la hoja de Excel
    for codigo in codigo_de_nomina:
        
        # Crea un nuevo diccionario para los datos del trabajador actual
        datos_trabajador = {}
        for row in hoja.iter_rows(min_row=6, min_col=1):
            
            # Comprueba si la celda contiene el código
            if row[0].value == codigo:

                # Obtiene los datos estáticos que no serán modificados por días
                codigo = hoja.cell(row=row[0].row, column=row[0].column).value
                salario = hoja.cell(row=row[0].row+3, column=row[0].column).value

                # Agrega los valores al diccionario de datos de captura del trabajador actual
                datos_trabajador['codigo'] = codigo
                datos_trabajador['salario'] = salario
                
                # Itera sobre los 7 días de la semana para agregar los valores correspondientes a cada día
                # Define una lista con los nombres de los días de la semana
                dias_semana = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]

                # Itera sobre los días de la semana para agregar los valores correspondientes a cada día
                for dia in dias_semana:
                    
                    obra = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+3).value
                    dia_trabajado = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+4).value
                    actividades = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+5).value
                    tem = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+6).value
                    tev = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+7).value
                    vacaciones = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+8).value
                    retardos = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+9).value
                    dia_festivo = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+10).value
                    velador = hoja.cell(row=row[0].row+1+dias_semana.index(dia), column=row[0].column+11).value
                    
                    # Agrega los valores correspondientes a cada día en el diccionario de datos del trabajador actual

                    if obra is not None:
                        datos_trabajador[dia + '_obra'] = obra
                    if dia_trabajado is not None:
                        datos_trabajador[dia + '_dia_trabajado'] = dia_trabajado
                    if actividades is not None:
                        datos_trabajador[dia + '_actividades'] = actividades
                    if tem is not None:
                        datos_trabajador[dia + '_tem'] = tem
                    if tev is not None:
                        datos_trabajador[dia + '_tev'] = tev
                    if vacaciones is not None:
                        datos_trabajador[dia + '_vacaciones'] = vacaciones
                    if retardos is not None:
                        datos_trabajador[dia + '_retardos'] = retardos
                    if dia_festivo is not None:
                        datos_trabajador[dia + '_dia_festivo'] = dia_festivo
                    if velador is not None:
                        datos_trabajador[dia + '_velador'] = velador

        # Agrega el diccionario de datos del trabajador actual a la lista de datos de captura del trabajador
        datos_de_captura_trabajador.append(datos_trabajador)

    return fecha_dia_semana, codigo_de_nomina, datos_de_captura_trabajador

fecha_dia_semana, codigo_de_nomina, datos_de_captura_trabajador = obtener_datos_de_captura(ruta_archivo_origen)
print(datos_de_captura_trabajador)

"""# GENERAMOS LA LISTA DE ACTIVIDADES
    # Creamos una sublista que representa las actividades de cada trabajador
    actividades = []
    for sublista in datos_de_captura:
        actividad = sublista[6]
        if actividad is None or actividad == "":
            actividad = "" # Si actividad es None o una cadena vacía, asignamos una cadena vacía
        actividades.append(actividad)

    lista_de_actividades = [[] for i in range(len(actividades))]

    for i, actividad in enumerate(actividades):
        # Dividir la cadena de texto en subcadenas de máximo 46 caracteres
        subcadenas = []
        while len(actividad) > 0:
            if len(actividad) <= 46:
                subcadenas.append(actividad)
                actividad = ""
            else:
                espacio = actividad.rfind(" ", 0, 46)
                if espacio == -1:
                    subcadenas.append(actividad[:46])
                    actividad = actividad[46:]
                else:
                    subcadenas.append(actividad[:espacio])
                    actividad = actividad[espacio+1:]

        # Agregar las subcadenas a la nueva lista correspondiente
        lista_de_actividades[i].extend(subcadenas)

    # GENERAMOS LA LISTA DE TRABAJADORES
    # La lista trabajadores, contendra las claves de cada uno de los trabajadores en datos_de_captura
    trabajadores = [lista[0] for lista in datos_de_captura]

    # Este ciclo for llenara la lista trabajador, enumerando a trabajadores iniciando desde el 0, para poder usarla como parametro en nuestra variable
    # Enumeramos los elementos de la lista y guardamos los índices en una lista
    trabajador = [i for i, _ in enumerate(trabajadores)]

    return datos_de_captura, lista_de_actividades, trabajador, fecha_domingo, fecha_dia_festivo

datos_de_captura, lista_de_actividades, trabajador, fecha_domingo, fecha_dia_festivo = obtener_datos_de_captura(ruta_archivo_origen)
"""


